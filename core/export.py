from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import OUTPUT_DIR


def _style(ws):
    fill = PatternFill("solid", fgColor="B22222")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 42)


def export_excel(summary: pd.DataFrame, details: pd.DataFrame, similarity: pd.DataFrame | None = None, output_dir: Path = OUTPUT_DIR) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"样品级分类汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    for idx, (title, df) in enumerate([
        ("样品级汇总", summary),
        ("重复谱明细", details),
        ("纯品相似度", similarity if similarity is not None else pd.DataFrame()),
    ]):
        ws = wb.active if idx == 0 else wb.create_sheet(title)
        ws.title = title
        if df.empty:
            ws.append(["说明"])
            ws.append(["暂无数据"])
        else:
            ws.append(list(df.columns))
            for row in df.itertuples(index=False):
                ws.append(list(row))
        _style(ws)
    wb.save(path)
    return path

